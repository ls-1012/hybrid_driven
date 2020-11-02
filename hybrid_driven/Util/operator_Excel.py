#!/usr/bin/env python
# -*- coding: utf-8 -*-
# -------------------------------------------------------------------------------
# File:         operator_Excel.py
# Author:       ls
# Date:         2020/10/12 17:56
# -------------------------------------------------------------------------------

from openpyxl import load_workbook
from openpyxl.styles import colors, Side, Border,PatternFill
from openpyxl.styles import Font
import os


class OPExcel:
    def __init__(self, filepath):
        """
        设置的行号和列号都是从现实来说的，都是从1开始，比如1,1表示第一行第一列单元格，即A1
        :param filepath: 文件路径
        """
        self.filepath = filepath
        # self.ws = None
        self.wb = None
        if not os.path.exists(filepath) or not (".xlsx" in filepath):
            print("输入的文件 %s 不存在或者文件类型不是.xlsx格式 " % self.filepath)
        else:
            self.wb = load_workbook(self.filepath)
            # 显示最后保存时的sheet页
            self.ws = self.wb.active

    # 已经知道sheetname,在设置一次self.sheet,意义在于显式切换sheet页
    def set_sheet_by_sheetname(self,sheetname):
        self.ws=self.wb[sheetname]
        # print("sheet obj:",self.ws)


    # 获取所有的sheet
    def get_sheetnames(self):
        return self.wb.sheetnames

    # 获得最大的行数
    def get_max_rows(self):
        return self.ws.max_row

    # 获得最大列数
    def get_max_cols(self):
        return self.ws.max_column

    # 通过下标获取对应sheet名，下标从1开始
    def set_sheet_by_index(self, index):
        if not isinstance(index, int):
            print("您设定的sheet序号 %s 不是整数，请重新设定" % index)
            return
        elif index > len(self.get_sheetnames()) or index <= 0:
            print("您设定的sheet序号 %s 不存在，请重新设定" % index)
            return
        else:
            sheet = self.get_sheetnames()[index - 1]
            self.ws=self.wb[sheet]
            return self.ws

    # 默认获取第一个sheet内容，可接受指定第几个sheet
    def getExcelContent(self):

        # print(self.ws[1][1].value)
        # print(self.wb.active)
        # print("self.sheet:  ",self.ws.title)
        res = []

        for row in self.ws.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            res.append(row_data)
        return res


    # 根据指定的行列写入单元格内容
    def writeContent(self, row, col, data):
        self.ws[row][col].value = data
        if data == "fail":
            self.ws[row][col].font = Font(color=colors.RED)
        self.wb.save(self.filepath)

    # 设置单元格样式
    def set_cell_style(self,rowNo=None,colNo=None,fontColor=None,bdColor=None,fgColor=None):
        if fontColor is None:
            font=Font(color=colors.BLACK)
        elif fontColor.lower()=="green":
            font = Font(color=colors.GREEN)
        else:
            font = Font(color=colors.RED)

        if bdColor is not None:
            bd = Side(style='thin', color="000000")
        else:
            bd = Side(style='thin', color=bdColor)

        if fgColor is None:
            fill=PatternFill(fill_type="solid",fgColor="DDDDDD")
        else:
            fill=PatternFill(fill_type="solid",fgColor=fgColor)

        # 这两句表明不管传不传行列号，说明操作整个单元格
        for row in self.ws.rows:
            for cell in row:
                cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        if colNo is None and rowNo is None:
            # rowNo=self.ws.rows
            # print("1")
            # for row in rowNo:
            #     for cell in row:
            #         cell.fill=fill
            pass
        elif rowNo  is not None and colNo is not None:

            self.ws.cell(rowNo,colNo).fill=fill


        elif rowNo is not None and colNo is None:

            for i in range(self.ws.max_column):
                self.ws.cell(rowNo,i+1).fill=fill
                # self.ws.cell(rowNo,i+1).value=55
                self.ws.cell(rowNo, i + 1).font=font


        elif colNo is not None and rowNo is None:

            for i in range(1,self.ws.max_row):
                self.ws.cell(i,colNo).fill=fill

        self.wb.save(self.filepath)

    def write_row_data(self,data):
        max_row = self.get_max_rows()
        if max_row==1 and self.getExcelContent()==[]:
            for i in range(len(data)):
                self.ws.cell(max_row, i + 1).value = data[i]
        else:
            for i in range(len(data)):
                if "fail" ==  data[i]:
                    # self.ws[max_row+1][i].font = Font(color=colors.RED)
                    # $$$上面的用法和下面的用法等价，但是需要注意的地方，用下标表示法列从0开始，用参数表示法，列从1开始，$$$
                    self.ws.cell(max_row+1,i+1).font=Font(color=colors.RED)
                    self.ws.cell(max_row+1,i+1).value=data[i]
                else:
                    self.ws.cell(max_row+1, i+1).value = data[i]
        # 下面这行会导致显示一空行
        # self.ws.append(data)
        self.wb.save(self.filepath)

    # 写入某一列的数据
    def write_col_data(self,col_no,data):
        for i in range(1,self.get_max_rows()):
            # self.ws[i+1][col_no-1].value=data
            self.writeContent(i+1,col_no-1,data)




    # 获取某一行的行对象
    def get_a_line(self,row_no):
        if not isinstance(row_no,int):
            print("输入的行号%s只能是整数" %row_no)
            return
        if not 0<row_no<self.ws.max_row+1:
            print("输入的行号%s超过行数范围" % row_no)
            return

        data=[]
        for row in self.ws.iter_rows():
            data.append(row)

        return data[row_no-1]



    # 获取某行的数据
    def get_line_data(self,row_no):
        # # 获取某一行的数据
        # return self.getExcelContent()[row_no-1]

        value=[]
        for cell in self.get_a_line(row_no):
            value.append(cell.value)

        return value

    # 获取指定列的值
    def get_col_data(self,col_no):
        if not isinstance(col_no, int):
            print("输入的列号%s只能是整数" % col_no)
            return
        if not 0 < col_no < self.ws.max_column + 1:
            print("输入的列号%s超过行数范围" % col_no)
            return

        data=[]
        for col in self.ws.iter_cols():
            tmp=[]
            for cell in col:
                tmp.append(cell.value)
            data.append(tmp)
        return data[col_no-1]

    # 获取每个单元格的值
    def get_cell_value(self,row_no,col_no):
        if not isinstance(row_no,int) or (not isinstance(col_no,int)):
            print("必须输入整数类型的行号或者列号")
            return

        if not 0<row_no<self.get_max_rows():
            print("输入的行号必须从1开始到最大行数之间")
            return
        if not 0<col_no<self.get_max_cols():
            print("输入的列号必须从1开始到最大列数之间")
            return

        return self.ws[row_no][col_no-1].value


    def create_sheet(self,sheetname):
        if sheetname not in self.get_sheetnames():
            self.wb.create_sheet(sheetname)
            self.wb.save(self.filepath)



if __name__ == "__main__":
    excel_data_file = r"C:\Users\LS\PycharmProjects\hybrid_driven\TestData\qq邮箱联系人.xlsx"
    wb = OPExcel(excel_data_file)
    #
    # wb.set_sheet_by_sheetname("联系人")
    # print(wb.getExcelContent())
    # wb.set_sheet_by_sheetname("qq账号")
    # print(wb.getExcelContent())
    #
    # print(wb.get_max_rows())
    # print(wb.get_line_data(1))
    # print(wb.get_col_data(2))
    # print(wb.get_cell_value(1,1))
    # wb.write_col_data(1,"*")
    # wb.set_cell_style(rowNo=1)
    if "xxxx"not in wb.get_sheetnames():
        wb.create_sheet("xxxx")
        print("创建成功")
    wb.set_sheet_by_sheetname("xxxx")
    wb.write_row_data([8,8,8,8,8])
    print(wb.getExcelContent())
    print(wb.get_max_rows())
    # wb.set_cell_style(wb.get_max_rows(),fgColor="BCEE68")
    wb.set_cell_style()